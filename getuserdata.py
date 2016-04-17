from openpyxl import load_workbook

def userdata(mobile):

    wb = load_workbook('userdata.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')

    for i in range(1,4):
        phone = str(sheet.cell(row=i, column=1).value)
        print phone
        if phone == str(mobile):
            return sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value

#print userdata(8439257665)

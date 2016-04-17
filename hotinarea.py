import math
from openpyxl import load_workbook

def getProducts(lat,lng):
    ans = []
    wb = load_workbook('productlocation.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(1,10):
        x = float(sheet.cell(row=i, column=1).value)
        y = float(sheet.cell(row=i, column=2).value)
        dlat = lat - x
        dlng = lng - y
        a = math.sin(dlat/2)**2 + (math.cos(lat) * math.cos(x) * math.sin(dlng/2)**2)
        c = 2*math.atan2(math.sqrt(a), math.sqrt(1-a))
        d = 6400*c
        print d
        if d < 50.00:
            myntra_url = sheet.cell(row=i, column=3).value
            ans.append(str(myntra_url))
    print ans
    return ans

getProducts(12.8925435,77.6395564)